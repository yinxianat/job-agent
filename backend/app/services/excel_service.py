"""
Excel export utilities.
Saves job search results and resume tracker to .xlsx files.
"""

import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="2563EB")   # brand blue
ALT_FILL     = PatternFill("solid", fgColor="EFF6FF")   # light blue
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER  = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header_row(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _style_data_rows(ws, num_rows: int, num_cols: int):
    for row in range(2, num_rows + 2):
        fill = ALT_FILL if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = THIN_BORDER


def jobs_to_excel(jobs: List[dict]) -> bytes:
    """Convert a list of job dicts to an in-memory Excel file (bytes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Results"

    headers = ["Title", "Company", "Location", "Posted Date", "Job URL", "Company URL", "Description"]
    ws.append(headers)

    for job in jobs:
        ws.append([
            job.get("title",       ""),
            job.get("company",     ""),
            job.get("location",    ""),
            job.get("posted_date", ""),
            job.get("job_url",     ""),
            job.get("company_url", ""),
            job.get("description", ""),
        ])

    _style_header_row(ws, len(headers))
    _style_data_rows(ws, len(jobs), len(headers))

    # Set column widths
    col_widths = [35, 25, 20, 15, 55, 45, 60]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def resume_tracker_to_excel(entries: List[dict]) -> bytes:
    """Build an Excel tracker for tailored resumes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume Tracker"

    headers = ["Resume Filename", "Job Title", "Company", "Location", "Job URL", "Company Description", "Created At"]
    ws.append(headers)

    for e in entries:
        ws.append([
            e.get("filename",            ""),
            e.get("job_title",           ""),
            e.get("company",             ""),
            e.get("location",            ""),
            e.get("job_url",             ""),
            e.get("company_description", ""),
            str(e.get("created_at",      "")),
        ])

    _style_header_row(ws, len(headers))
    _style_data_rows(ws, len(entries), len(headers))

    col_widths = [40, 30, 25, 20, 55, 60, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def batch_tracker_to_excel(entries: List[dict]) -> bytes:
    """Excel tracker for a batch generation run — includes cover letter filenames."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Tracker"

    headers = [
        "Resume Filename", "Cover Letter Filename", "Job Title", "Company",
        "Location", "Job URL", "Company Description", "Status",
    ]
    ws.append(headers)

    STATUS_COLOR = {
        "done":       "D1FAE5",
        "error":      "FEE2E2",
        "pending":    "FEF9C3",
        "processing": "DBEAFE",
    }

    for row_idx, e in enumerate(entries, start=2):
        ws.append([
            e.get("filename",              ""),
            e.get("cover_letter_filename", ""),
            e.get("job_title",             ""),
            e.get("company",               ""),
            e.get("location",              ""),
            e.get("job_url",               ""),
            e.get("company_description",   ""),
            e.get("status",                ""),
        ])
        status = e.get("status", "")
        color  = STATUS_COLOR.get(status, "FFFFFF")
        ws.cell(row=row_idx, column=8).fill = PatternFill("solid", fgColor=color)

    _style_header_row(ws, len(headers))

    for row in range(2, len(entries) + 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = THIN_BORDER

    col_widths = [40, 42, 28, 25, 20, 55, 60, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
